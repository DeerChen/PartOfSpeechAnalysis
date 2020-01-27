'''
@Author: Senkita
'''

import re
import json
import time
import pandas as pd
from textblob import TextBlob
from openpyxl import Workbook

pd.set_option('mode.chained_assignment', None)

class Processing:
    def __init__(self, file_path, config_file):
        with open(file_path, 'r', encoding='utf-8-sig') as f:
            self.text = ''.join(list(filter(self.cn_condition, f.read())))
        self.doc = TextBlob(self.text)

        with open(config_file, 'r', encoding='utf-8-sig') as f:
            self.part_of_speech = json.loads(f.read())

    # 双字节字符判断
    @staticmethod
    def cn_condition(string):
        cn_pattern = re.compile(r'[\x00-\xff]')
        cn_match = re.findall(cn_pattern, string)
        
        return len(cn_match) != 0
    
    # 专有判断
    @staticmethod
    def proper_condition(string):
        proper_pattern = re.compile(r'^专有')
        proper_match = re.findall(proper_pattern, string)
        
        return len(proper_match) != 0

    # 词性标注
    def analysis(self):
        result = {}
        for v in self.part_of_speech.values():
            result[v] = {}

        for i in self.doc.tags:
            translation = self.part_of_speech[i[1]]
            if self.proper_condition(translation):
                word = i[0]
            else:
                word = i[0].lower()

            if result[translation].get(word) != None:
                result[translation][word] += 1
            else:
                result[translation][word] = 1

        return result

    # 保存为Excel
    def output(self, result):
        name = '{}.xlsx'.format(int(time.time()))
        wb = Workbook()
        ws = wb.active

        num = 0
        for i in result.keys():
            ws.cell(1, num * 2 + 1).value = i
            ws.merge_cells(
                start_row=1,
                start_column=num * 2 + 1,
                end_row=1,
                end_column=num * 2 + 2
            )
            ws.cell(2, num * 2 + 1).value = '单词'
            ws.cell(2, num * 2 + 2).value = '频次'

            if result[i] != {}:
                data = pd.DataFrame(result[i], index=['频次']).T
                for j in range(len(data)):
                    word = data.index[j]
                    frequency = int(data['频次'][j])

                    ws.cell(j + 3, num * 2 + 1).value = word
                    ws.cell(j + 3, num * 2 +
                            2).value = '{}'.format(frequency)
            num += 1

        wb.save(name)
        return name